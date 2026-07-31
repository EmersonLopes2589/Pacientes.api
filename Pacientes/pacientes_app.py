"""
Consulta de Pacientes — Excel -> SQLite -> Interface gráfica (Tkinter)

O que este script faz:
    1. Lê as colunas Atendimento / Data Atendimento / Prontuário / Nome / Data Registro
       de uma planilha .xlsx.
    2. Grava (ou atualiza) esses dados num banco SQLite local (pacientes.db).
    3. Abre uma janela com um campo de busca por nome e uma tabela com os resultados.

Requisitos:
    pip install openpyxl
    (tkinter já vem junto com o instalador padrão do Python no Windows;
     no Linux, se faltar, instale com: sudo apt install python3-tk)

Como usar:
    python pacientes_app.py
    python pacientes_app.py "C:\\caminho\\para\\Relacao_Pacientes_Organizado.xlsx"

Se você não informar o caminho e o arquivo padrão não for encontrado na
pasta atual, uma janela para escolher o arquivo será aberta automaticamente.
"""

import os
import sys
import sqlite3
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl

# ---------------------------------------------------------------------------
# Configurações
# ---------------------------------------------------------------------------

DB_PATH = "pacientes.db"
DEFAULT_XLSX = "Relacao_Pacientes_Organizado.xlsx"

# Nomes das colunas exatamente como aparecem no cabeçalho da planilha
COLUNAS_DESEJADAS = ["Atendimento", "Data Atendimento", "Prontuário", "Nome", "Data Registro"]


# ---------------------------------------------------------------------------
# Banco de dados
# ---------------------------------------------------------------------------

def criar_banco(db_path=DB_PATH):
    """Cria (se não existir) a tabela 'pacientes' e devolve a conexão."""
    conn = sqlite3.connect(db_path)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pacientes (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            atendimento       INTEGER,
            data_atendimento  TEXT,
            prontuario        INTEGER,
            nome              TEXT,
            data_registro     TEXT
        )
    """)
    conn.commit()
    return conn


def formatar_data(valor):
    """Converte a célula de data (datetime ou texto) para 'DD/MM/AAAA' ou None."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    return str(valor).strip()


def importar_xlsx(xlsx_path, conn):
    """Lê a planilha e (re)popula a tabela 'pacientes' no SQLite."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    cabecalho = [c.value for c in ws[1]]
    indices = {}
    for nome_coluna in COLUNAS_DESEJADAS:
        if nome_coluna not in cabecalho:
            raise ValueError(
                f"Coluna '{nome_coluna}' não encontrada na planilha.\n"
                f"Colunas disponíveis: {cabecalho}"
            )
        indices[nome_coluna] = cabecalho.index(nome_coluna)

    cur = conn.cursor()
    cur.execute("DELETE FROM pacientes")  # evita duplicar dados a cada reimportação

    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        atendimento = row[indices["Atendimento"]]
        data_atendimento = formatar_data(row[indices["Data Atendimento"]])
        prontuario = row[indices["Prontuário"]]
        nome = row[indices["Nome"]]
        data_registro = formatar_data(row[indices["Data Registro"]])

        if atendimento is None and nome is None:
            continue  # pula linhas em branco

        cur.execute(
            """
            INSERT INTO pacientes (atendimento, data_atendimento, prontuario, nome, data_registro)
            VALUES (?, ?, ?, ?, ?)
            """,
            (atendimento, data_atendimento, prontuario, nome, data_registro),
        )
        total += 1

    conn.commit()
    return total


def buscar_pacientes(conn, termo_nome=""):
    """Retorna os registros cujo nome contém o termo buscado (case-insensitive)."""
    cur = conn.cursor()
    if termo_nome:
        cur.execute(
            """
            SELECT atendimento, data_atendimento, prontuario, nome, data_registro
            FROM pacientes
            WHERE nome LIKE ?
            ORDER BY nome
            """,
            (f"%{termo_nome}%",),
        )
    else:
        cur.execute(
            """
            SELECT atendimento, data_atendimento, prontuario, nome, data_registro
            FROM pacientes
            ORDER BY nome
            """
        )
    return cur.fetchall()


# ---------------------------------------------------------------------------
# Interface gráfica
# ---------------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self, conn):
        super().__init__()
        self.conn = conn
        self.title("Consulta de Pacientes")
        self.geometry("900x520")
        self.minsize(700, 400)

        self._montar_interface()
        self._atualizar_tabela()

    def _montar_interface(self):
        # --- Barra de busca ---
        topo = ttk.Frame(self, padding=10)
        topo.pack(fill="x")

        ttk.Label(topo, text="Buscar por nome:").pack(side="left")

        self.busca_var = tk.StringVar()
        entrada = ttk.Entry(topo, textvariable=self.busca_var, width=40)
        entrada.pack(side="left", padx=8)
        entrada.bind("<KeyRelease>", lambda e: self._atualizar_tabela())
        entrada.focus()

        ttk.Button(topo, text="Limpar", command=self._limpar_busca).pack(side="left")

        # --- Tabela de resultados ---
        area_tabela = ttk.Frame(self)
        area_tabela.pack(fill="both", expand=True, padx=10, pady=(0, 5))

        colunas = ("atendimento", "data_atendimento", "prontuario", "nome", "data_registro")
        titulos = ("Atendimento", "Data Atendimento", "Prontuário", "Nome", "Data Registro")

        self.tabela = ttk.Treeview(area_tabela, columns=colunas, show="headings")
        for col, titulo in zip(colunas, titulos):
            self.tabela.heading(col, text=titulo)
            largura = 260 if col == "nome" else 140
            ancora = "w" if col == "nome" else "center"
            self.tabela.column(col, width=largura, anchor=ancora)

        scrollbar = ttk.Scrollbar(area_tabela, orient="vertical", command=self.tabela.yview)
        self.tabela.configure(yscrollcommand=scrollbar.set)

        self.tabela.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # --- Barra de status ---
        self.status_var = tk.StringVar()
        ttk.Label(self, textvariable=self.status_var, padding=(10, 4)).pack(fill="x")

    def _limpar_busca(self):
        self.busca_var.set("")
        self._atualizar_tabela()

    def _atualizar_tabela(self):
        termo = self.busca_var.get().strip()

        for item in self.tabela.get_children():
            self.tabela.delete(item)

        resultados = buscar_pacientes(self.conn, termo)
        for linha in resultados:
            self.tabela.insert("", "end", values=linha)

        self.status_var.set(f"{len(resultados)} paciente(s) encontrado(s)")


# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------

def escolher_planilha():
    """Abre um seletor de arquivo para o usuário escolher a planilha manualmente."""
    root = tk.Tk()
    root.withdraw()
    caminho = filedialog.askopenfilename(
        title="Selecione a planilha de pacientes",
        filetypes=[("Planilhas Excel", "*.xlsx")],
    )
    root.destroy()
    return caminho


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX

    if not os.path.exists(xlsx_path):
        print(f"Arquivo '{xlsx_path}' não encontrado na pasta atual.")
        xlsx_path = escolher_planilha()
        if not xlsx_path:
            print("Nenhum arquivo selecionado. Encerrando.")
            return

    conn = criar_banco(DB_PATH)

    try:
        total = importar_xlsx(xlsx_path, conn)
        print(f"{total} registro(s) importado(s) para '{DB_PATH}'.")
    except Exception as e:
        messagebox.showerror("Erro ao importar planilha", str(e))
        conn.close()
        return

    app = App(conn)
    app.mainloop()
    conn.close()


if __name__ == "__main__":
    main()
